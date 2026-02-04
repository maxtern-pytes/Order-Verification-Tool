import os
import json
import csv
import io
import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor
from flask import Flask, request, jsonify, render_template, redirect, url_for, Response, send_file
from flask_basicauth import BasicAuth
from datetime import datetime
import pytz
from dotenv import load_dotenv

# Load .env for local development
load_dotenv(override=True)

app = Flask(__name__)
IST = pytz.timezone('Asia/Kolkata')

# --- Configuration ---
DATABASE_URL = os.getenv("DATABASE_URL")
if DATABASE_URL:
    DATABASE_URL = DATABASE_URL.strip()
    print(f"DEBUG: Connecting to DB... Ending with: {DATABASE_URL[-20:]}")
else:
    print("WARNING: DATABASE_URL not found. App will crash if database is accessed.")

app.config['BASIC_AUTH_USERNAME'] = os.getenv("BASIC_AUTH_USERNAME", "admin")
app.config['BASIC_AUTH_PASSWORD'] = os.getenv("BASIC_AUTH_PASSWORD", "admin123")

# Viewer credentials (read-only access)
VIEWER_USERNAME = os.getenv("VIEWER_USERNAME", "viewer")
VIEWER_PASSWORD = os.getenv("VIEWER_PASSWORD", "viewer123")

basic_auth = BasicAuth(app)

# Custom auth check for viewer
def check_viewer_auth(username, password):
    return username == VIEWER_USERNAME and password == VIEWER_PASSWORD

# --- Database Setup ---
def get_db_connection():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

# --- Customer Management Functions ---
def create_or_update_customer(order):
    """Create or update customer profile from order data"""
    try:
        conn = get_db_connection()
        c = conn.cursor()
        
        phone = order.get('phone')
        if not phone:
            return
        
        # Check if customer exists
        c.execute('SELECT * FROM customers WHERE phone = %s', (phone,))
        existing = c.fetchone()
        
        if existing:
            # Update existing customer
            c.execute('''
                UPDATE customers SET
                    name = COALESCE(%s, name),
                    email = COALESCE(%s, email),
                    last_order_date = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE phone = %s
            ''', (order.get('customer_name'), order.get('email'), order.get('timestamp'), phone))
        else:
            # Create new customer
            c.execute('''
                INSERT INTO customers (phone, name, email, first_order_date, last_order_date, tags)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (phone, order.get('customer_name'), order.get('email'), 
                  order.get('timestamp'), order.get('timestamp'), '["New Customer"]'))
        
        conn.commit()
        conn.close()
        
        # Update customer stats
        update_customer_stats(phone)
    except Exception as e:
        print(f"Error creating/updating customer: {e}")

def update_customer_stats(phone):
    """Recalculate customer statistics"""
    try:
        conn = get_db_connection()
        c = conn.cursor()
        
        # Get order statistics
        c.execute('''
            SELECT 
                COUNT(*) as total_orders,
                COUNT(*) FILTER (WHERE status = 'Confirmed') as confirmed_orders,
                COUNT(*) FILTER (WHERE status = 'Cancelled') as cancelled_orders,
                COALESCE(SUM(CASE 
                    WHEN status = 'Confirmed' AND total IS NOT NULL AND total != '' 
                    THEN CAST(NULLIF(REGEXP_REPLACE(total, '[^0-9.]', '', 'g'), '') AS DECIMAL(10,2))
                    ELSE 0 
                END), 0) as total_spent,
                COUNT(*) FILTER (WHERE rto_risk = 'High') as rto_count,
                json_agg(DISTINCT address) FILTER (WHERE address IS NOT NULL AND address != '') as addresses,
                json_agg(DISTINCT state) FILTER (WHERE state IS NOT NULL AND state != '') as states
            FROM orders
            WHERE phone = %s
        ''', (phone,))
        
        stats = c.fetchone()
        
        # Get preferred payment method
        c.execute('''
            SELECT payment_method
            FROM orders
            WHERE phone = %s AND payment_method IS NOT NULL
            GROUP BY payment_method
            ORDER BY COUNT(*) DESC
            LIMIT 1
        ''', (phone,))
        preferred_payment = c.fetchone()
        
        # Get preferred delivery type
        c.execute('''
            SELECT delivery_type
            FROM orders
            WHERE phone = %s AND delivery_type IS NOT NULL
            GROUP BY delivery_type
            ORDER BY COUNT(*) DESC
            LIMIT 1
        ''', (phone,))
        preferred_delivery = c.fetchone()
        
        # Auto-tag based on stats
        tags = []
        if stats['total_spent'] > 10000:
            tags.extend(['VIP', 'High Value'])
        if stats['total_orders'] >= 5:
            tags.append('Frequent Buyer')
        if stats['cancelled_orders'] > 2:
            tags.append('High Risk')
        if stats['total_orders'] == 1:
            tags.append('New Customer')
        if stats['confirmed_orders'] >= 3:
            tags.append('Loyal')
        
        # Update customer
        c.execute('''
            UPDATE customers SET
                total_orders = %s,
                confirmed_orders = %s,
                cancelled_orders = %s,
                total_spent = %s,
                rto_count = %s,
                addresses = %s,
                states = %s,
                preferred_payment = %s,
                preferred_delivery = %s,
                tags = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE phone = %s
        ''', (
            stats['total_orders'],
            stats['confirmed_orders'],
            stats['cancelled_orders'],
            stats['total_spent'],
            stats['rto_count'],
            json.dumps(stats['addresses']) if stats['addresses'] else '[]',
            json.dumps(stats['states']) if stats['states'] else '[]',
            preferred_payment['payment_method'] if preferred_payment else None,
            preferred_delivery['delivery_type'] if preferred_delivery else None,
            json.dumps(tags),
            phone
        ))
        
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error updating customer stats: {e}")

def get_customer_by_phone(phone):
    """Get customer profile by phone number"""
    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('SELECT * FROM customers WHERE phone = %s', (phone,))
        customer = c.fetchone()
        conn.close()
        return customer
    except Exception as e:
        print(f"Error getting customer: {e}")
        return None

def get_all_customers(search=None, filter_type=None, sort_by='last_order_date', sort_order='DESC'):
    """Get all customers with optional filtering and sorting"""
    try:
        conn = get_db_connection()
        c = conn.cursor()
        
        query = 'SELECT * FROM customers WHERE 1=1'
        params = []
        
        # Search filter
        if search:
            query += ' AND (name ILIKE %s OR phone ILIKE %s OR email ILIKE %s)'
            search_pattern = f'%{search}%'
            params.extend([search_pattern, search_pattern, search_pattern])
        
        # Type filters
        if filter_type == 'repeat':
            query += ' AND total_orders > 1'
        elif filter_type == 'vip':
            query += ' AND total_spent > 10000'
        elif filter_type == 'high_risk':
            query += ' AND cancelled_orders > 2'
        elif filter_type == 'new':
            query += ' AND total_orders = 1'
        
        # Sorting
        valid_sorts = ['total_orders', 'total_spent', 'last_order_date', 'name']
        if sort_by in valid_sorts:
            query += f' ORDER BY {sort_by} {sort_order}'
        
        c.execute(query, params)
        customers = c.fetchall()
        conn.close()
        return customers
    except Exception as e:
        print(f"Error getting customers: {e}")
        return []

def init_db():
    if not DATABASE_URL:
        return
    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('''
            CREATE TABLE IF NOT EXISTS orders (
                id TEXT PRIMARY KEY,
                customer_name TEXT,
                email TEXT,
                phone TEXT,
                address TEXT,
                source TEXT,
                products TEXT,  -- JSON string
                total TEXT,
                status TEXT,
                timestamp TEXT,
                notes TEXT,
                delivery_type TEXT,
                state TEXT,
                payment_method TEXT
            )
        ''')
        
        # Migrations
        c.execute("SELECT column_name FROM information_schema.columns WHERE table_name='orders'")
        existing_columns = [row[0] for row in c.fetchall()]
        
        # Ensure all columns are present, adding if missing
        if 'address' not in existing_columns:
            c.execute("ALTER TABLE orders ADD COLUMN address TEXT")
        if 'notes' not in existing_columns:
            c.execute("ALTER TABLE orders ADD COLUMN notes TEXT")
        if 'delivery_type' not in existing_columns:
            c.execute("ALTER TABLE orders ADD COLUMN delivery_type TEXT DEFAULT 'Standard'")
        if 'state' not in existing_columns:
            c.execute("ALTER TABLE orders ADD COLUMN state TEXT")
        if 'payment_method' not in existing_columns:
            c.execute("ALTER TABLE orders ADD COLUMN payment_method TEXT DEFAULT 'Prepaid'")
        if 'email' not in existing_columns:
            c.execute("ALTER TABLE orders ADD COLUMN email TEXT")

        conn.commit()
        conn.close()
        print("Database initialized successfully.")
    except Exception as e:
        print(f"Error initializing database: {e}")

init_db()

# --- Helper Functions ---

def calculate_rto_risk(payment_method, state):
    """Calculate RTO risk for Shopify orders based on payment method and state"""
    # COD orders have higher RTO risk
    if payment_method == "COD":
        # High-risk states for COD
        high_risk_states = ["Bihar", "Jharkhand", "Uttar Pradesh", "West Bengal", "Odisha", "Assam"]
        if state in high_risk_states:
            return "HIGH"
        return "MEDIUM"
    # Prepaid orders are generally low risk
    return "LOW"

def normalize_shopify_order(data):
    try:
        products = [f"{item.get('name')} (Qty: {item.get('quantity', 1)})" for item in data.get("line_items", [])]
        shipping = data.get("shipping_address", {})
        address = f"{shipping.get('address1', '')}, {shipping.get('city', '')}, {shipping.get('zip', '')}"
        delivery_type = "Standard"
        if "Express" in data.get("tags", ""):
            delivery_type = "Express"
        
        # Payment Method - check gateway and payment_gateway_names
        gateway = str(data.get("gateway", "")).lower()
        payment_gateways = data.get("payment_gateway_names", [])
        payment_method = "Prepaid"
        
        if "cod" in gateway or "cash" in gateway:
            payment_method = "COD"
        elif payment_gateways:
            gateway_str = str(payment_gateways).lower()
            if "cod" in gateway_str or "cash" in gateway_str:
                payment_method = "COD"
        
        state = shipping.get("province", "")
        rto_risk = calculate_rto_risk(payment_method, state)
            
        return {
            "id": str(data.get("name", "N/A")),
            "customer_name": f"{data.get('customer', {}).get('first_name', '')} {data.get('customer', {}).get('last_name', '')}".strip() or "Guest",
            "email": data.get('customer', {}).get('email', ''),
            "phone": shipping.get("phone") or data.get("customer", {}).get("phone") or "No Phone",
            "address": address,
            "state": state,
            "payment_method": payment_method,
            "rto_risk": rto_risk,
            "source": "Shopify",
            "products": json.dumps(products),
            "total": data.get("total_price", "0.00"),
            "status": "Pending",
            "timestamp": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
            "notes": "",
            "delivery_type": delivery_type
        }
    except Exception as e:
        print(f"Error parsing Shopify data: {e}")
        return None

def normalize_shiprocket_order(data):
    try:
        products = [f"{item.get('name')} (Qty: {item.get('quantity', 1)})" for item in data.get("products", [])]
        address = f"{data.get('shipping_address', '')}, {data.get('shipping_city', '')}, {data.get('shipping_pincode', '')}"
        
        # Check tags for Express
        raw_tags = data.get("tags") or data.get("order_tags") or "" 
        if isinstance(raw_tags, list):
            tags_str = ",".join(raw_tags)
        else:
            tags_str = str(raw_tags)
            
        delivery_type = "Express" if "Express" in tags_str else "Standard"
        
        # Payment Method - Check multiple possible fields
        payment_method = "Prepaid"  # Default
        
        # Check payment_method field
        payment_raw = str(data.get("payment_method", "")).upper()
        if "COD" in payment_raw or "CASH" in payment_raw:
            payment_method = "COD"
        
        # Also check cod field if present
        if data.get("cod") == 1 or data.get("cod") == "1" or data.get("is_cod"):
            payment_method = "COD"
            
        # Check payment_gateway field
        payment_gateway = str(data.get("payment_gateway", "")).upper()
        if "COD" in payment_gateway or "CASH" in payment_gateway:
            payment_method = "COD"
        
        # Extract RTO risk from tags (Shiprocket provides this)
        rto_risk = "LOW"  # Default
        if "HIGH RISK" in tags_str.upper() or "HIGH_RISK" in tags_str.upper():
            rto_risk = "HIGH"
        elif "MEDIUM RISK" in tags_str.upper() or "MEDIUM_RISK" in tags_str.upper():
            rto_risk = "MEDIUM"

        return {
            "id": str(data.get("channel_order_id") or data.get("order_id", "N/A")),
            "customer_name": data.get("customer_name", "Guest"),
            "email": data.get("customer_email", ""),
            "phone": data.get("customer_phone", "No Phone"),
            "address": address,
            "state": data.get("shipping_state", ""),
            "payment_method": payment_method,
            "rto_risk": rto_risk,
            "source": "Shiprocket",
            "products": json.dumps(products),
            "total": data.get("net_total", "0.00"),
            "status": "Pending",
            "timestamp": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
            "notes": "",
            "delivery_type": delivery_type
        }
    except Exception as e:
        print(f"Error parsing Shiprocket data: {e}")
        return None

def save_order(order):
    conn = get_db_connection()
    c = conn.cursor(cursor_factory=RealDictCursor)
    
    # Check for existing data to preserve
    c.execute('SELECT notes, delivery_type, state, payment_method, email, rto_risk FROM orders WHERE id = %s', (order['id'],))
    existing = c.fetchone()
    notes = existing['notes'] if existing and existing.get('notes') else order.get('notes', '')
    delivery_type = existing['delivery_type'] if existing and existing.get('delivery_type') else order.get('delivery_type', 'Standard')
    state = order.get('state') or (existing['state'] if existing else '')
    payment_method = order.get('payment_method') or (existing['payment_method'] if existing else 'Prepaid')
    email = order.get('email') or (existing['email'] if existing else '')
    rto_risk = order.get('rto_risk') or (existing['rto_risk'] if existing else 'LOW')

    # Postgres UPSERT
    query = '''
        INSERT INTO orders (id, customer_name, email, phone, address, source, products, total, status, timestamp, notes, delivery_type, state, payment_method, rto_risk)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (id) DO UPDATE SET
            customer_name = EXCLUDED.customer_name,
            email = EXCLUDED.email,
            phone = EXCLUDED.phone,
            address = EXCLUDED.address,
            source = EXCLUDED.source,
            products = EXCLUDED.products,
            total = EXCLUDED.total,
            status = EXCLUDED.status,
            timestamp = EXCLUDED.timestamp,
            notes = %s,
            delivery_type = %s,
            state = %s,
            payment_method = %s,
            rto_risk = %s
    '''
    c.execute(query, (
        order['id'], order['customer_name'], email, order['phone'], order.get('address', ''), order['source'], 
        order['products'], order['total'], order['status'], order['timestamp'], notes, delivery_type, state, payment_method, rto_risk,
        notes, delivery_type, state, payment_method, rto_risk
    ))
    conn.commit()
    conn.close()
    
    # Auto-create/update customer profile
    create_or_update_customer(order)

def get_orders(status_filter='Pending', start_date=None, end_date=None, search_query=None, payment_filter=None, delivery_filter=None, state_filter=None):
    conn = get_db_connection()
    query = 'SELECT * FROM orders WHERE status = %s'
    params = [status_filter]

    if start_date:
        query += ' AND timestamp >= %s'
        params.append(start_date + " 00:00:00")
    if end_date:
        query += ' AND timestamp <= %s'
        params.append(end_date + " 23:59:59")
    
    if search_query:
        # Smart search: if 5 or fewer digits, search only Order ID
        # Otherwise search phone, name, email, address
        if search_query.isdigit() and len(search_query) <= 5:
            query += ' AND id ILIKE %s'
            wildcard = f"%{search_query}%"
            params.append(wildcard)
        else:
            query += ' AND (customer_name ILIKE %s OR phone ILIKE %s OR email ILIKE %s OR address ILIKE %s)'
            wildcard = f"%{search_query}%"
            params.extend([wildcard, wildcard, wildcard, wildcard])
    
    # New filters
    if payment_filter:
        query += ' AND payment_method = %s'
        params.append(payment_filter)
    
    if delivery_filter:
        query += ' AND delivery_type = %s'
        params.append(delivery_filter)
    
    if state_filter:
        query += ' AND state = %s'
        params.append(state_filter)
    
    query += ' ORDER BY timestamp DESC'
    c = conn.cursor()
    c.execute(query, params)
    orders = c.fetchall()
    
    orders_list = []
    for row in orders:
        order = dict(row)
        try:
            order['products'] = json.loads(order['products'])
        except:
            order['products'] = []
        
        # Set default customer values
        order['is_repeat_customer'] = False
        order['customer_total_orders'] = 0
        order['customer_total_spent'] = 0
        order['customer_tags'] = '[]'
            
        orders_list.append(order)
    
    # Batch fetch customer data for all orders with phones
    if orders_list:
        phones = [o['phone'] for o in orders_list if o.get('phone')]
        if phones:
            try:
                # Use a single query to get all customer data
                placeholders = ','.join(['%s'] * len(phones))
                c.execute(f'''
                    SELECT phone, total_orders, total_spent, tags
                    FROM customers
                    WHERE phone IN ({placeholders})
                ''', phones)
                customers_dict = {row['phone']: row for row in c.fetchall()}
                
                # Enrich orders with customer data
                for order in orders_list:
                    if order.get('phone') and order['phone'] in customers_dict:
                        customer = customers_dict[order['phone']]
                        order['customer_total_orders'] = customer['total_orders']
                        order['customer_total_spent'] = customer['total_spent']
                        order['is_repeat_customer'] = customer['total_orders'] > 1
                        order['customer_tags'] = customer.get('tags', '[]')
            except Exception as e:
                # If customers table doesn't exist yet, just skip customer enrichment
                print(f"Customer enrichment skipped: {e}")
    
    conn.close()
    return orders_list

def get_daily_summary():
    conn = get_db_connection()
    # Postgres substring syntax: substring(string from start for length)
    # timestamp is TEXT in our schema, so substring works.
    query = '''
        SELECT substring(timestamp, 1, 10) as day, 
               COUNT(*) as total,
               SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
               SUM(CASE WHEN status = 'Confirmed' THEN 1 ELSE 0 END) as confirmed,
               SUM(CASE WHEN status = 'Cancelled' THEN 1 ELSE 0 END) as cancelled,
               SUM(CASE WHEN status = 'Call Again' THEN 1 ELSE 0 END) as call_again
        FROM orders
        GROUP BY day
        ORDER BY day DESC
    '''
    c = conn.cursor()
    c.execute(query)
    rows = c.fetchall()
    conn.close()
    return rows

# --- Routes (Protected) ---

@app.route('/')
@basic_auth.required
def dashboard():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    payment = request.args.get('payment')
    delivery = request.args.get('delivery')
    state = request.args.get('state')
    try:
        orders = get_orders('Pending', start_date, end_date, search, payment, delivery, state)
    except Exception as e:
        return f"Database Error: {e}. Did you set DATABASE_URL in .env?", 500
    return render_template('dashboard.html', orders=orders, view='Pending', start_date=start_date, end_date=end_date, search=search)

@app.route('/call-again')
@basic_auth.required
def call_again_page():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    payment = request.args.get('payment')
    delivery = request.args.get('delivery')
    state = request.args.get('state')
    orders = get_orders('Call Again', start_date, end_date, search, payment, delivery, state)
    return render_template('dashboard.html', orders=orders, view='Call Again', start_date=start_date, end_date=end_date, search=search)

@app.route('/reports')
@basic_auth.required
def reports_page():
    summary = get_daily_summary()
    return render_template('dashboard.html', orders=[], view='Reports', summary=summary)

@app.route('/update_status', methods=['POST'])
@basic_auth.required
def update_status():
    order_id = request.form['order_id']
    new_status = request.form['status']
    conn = get_db_connection()
    c = conn.cursor()
    
    # Get order phone before update
    c.execute('SELECT phone FROM orders WHERE id = %s', (order_id,))
    order = c.fetchone()
    
    c.execute('UPDATE orders SET status = %s WHERE id = %s', (new_status, order_id))
    conn.commit()
    conn.close()
    
    # Update customer stats
    if order and order['phone']:
        update_customer_stats(order['phone'])
    
    return redirect(request.referrer or '/')

@app.route('/bulk_delete', methods=['POST'])
@basic_auth.required
def bulk_delete():
    """Delete multiple orders by their IDs"""
    try:
        data = request.get_json()
        order_ids = data.get('order_ids', [])
        
        if not order_ids:
            return jsonify({'success': False, 'error': 'No orders selected'}), 400
        
        conn = get_db_connection()
        c = conn.cursor()
        
        # Delete orders with matching IDs
        placeholders = ','.join(['%s'] * len(order_ids))
        query = f'DELETE FROM orders WHERE id IN ({placeholders})'
        c.execute(query, order_ids)
        
        conn.commit()
        deleted_count = c.rowcount
        conn.close()
        
        return jsonify({'success': True, 'deleted': deleted_count})
    except Exception as e:
        print(f"Error in bulk_delete: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/clear_all', methods=['POST'])
@basic_auth.required
def clear_all():
    """Delete all orders in a specific view/status"""
    try:
        view = request.args.get('view', 'Confirmed')
        
        conn = get_db_connection()
        c = conn.cursor()
        
        # Delete all orders with the specified status
        c.execute('DELETE FROM orders WHERE status = %s', (view,))
        
        conn.commit()
        deleted_count = c.rowcount
        conn.close()
        
        return jsonify({'success': True, 'deleted': deleted_count})
    except Exception as e:
        print(f"Error in clear_all: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/confirmed')
@basic_auth.required
def confirmed_page():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    payment = request.args.get('payment')
    delivery = request.args.get('delivery')
    state = request.args.get('state')
    orders = get_orders('Confirmed', start_date, end_date, search, payment, delivery, state)
    return render_template('dashboard.html', orders=orders, view='Confirmed', start_date=start_date, end_date=end_date, search=search)

@app.route('/cancelled')
@basic_auth.required
def cancelled_page():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    payment = request.args.get('payment')
    delivery = request.args.get('delivery')
    state = request.args.get('state')
    orders = get_orders('Cancelled', start_date, end_date, search, payment, delivery, state)
    return render_template('dashboard.html', orders=orders, view='Cancelled', start_date=start_date, end_date=end_date, search=search)

@app.route('/viewer')
def viewer_dashboard():
    """Read-only dashboard for viewing confirmed orders only"""
    print("=" * 80)
    print("VIEWER DASHBOARD ACCESS ATTEMPT")
    print(f"Expected Username: {VIEWER_USERNAME}")
    print(f"Expected Password: {VIEWER_PASSWORD}")
    
    auth = request.authorization
    if auth:
        print(f"Received Username: {auth.username}")
        print(f"Password Match: {auth.password == VIEWER_PASSWORD}")
    else:
        print("No authorization header received")
    print("=" * 80)
    
    if not auth or not check_viewer_auth(auth.username, auth.password):
        return Response(
            'Viewer login required', 401,
            {'WWW-Authenticate': 'Basic realm="Viewer Login"'}
        )
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search')
    payment = request.args.get('payment')
    delivery = request.args.get('delivery')
    state = request.args.get('state')
    orders = get_orders('Confirmed', start_date, end_date, search, payment, delivery, state)
    return render_template('viewer.html', orders=orders, start_date=start_date, end_date=end_date, search=search)

@app.route('/update_order_details/<order_id>', methods=['POST'])
@basic_auth.required
def update_order_details(order_id):
    new_products_text = request.form.get('products_text')
    new_address = request.form.get('address')
    new_phone = request.form.get('phone')
    new_notes = request.form.get('notes')
    new_delivery = request.form.get('delivery_type') or 'Standard'
    
    try:
        products_list = json.loads(new_products_text)
    except:
        products_list = [p.strip() for p in new_products_text.split(',') if p.strip()]

    conn = get_db_connection()
    c = conn.cursor()
    c.execute('UPDATE orders SET products = %s, address = %s, phone = %s, notes = %s, delivery_type = %s WHERE id = %s', 
                 (json.dumps(products_list), new_address, new_phone, new_notes, new_delivery, order_id))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('dashboard'))

# --- Exports (Protected) ---

def get_orders_for_export(start_date, end_date, status=None, delivery_type=None):
    conn = get_db_connection()
    c = conn.cursor()
    query = "SELECT * FROM orders WHERE 1=1"
    params = []
    
    if start_date:
        query += " AND date(timestamp) >= %s"
        params.append(start_date)
    if end_date:
        query += " AND date(timestamp) <= %s"
        params.append(end_date)
    if status:
        query += " AND status = %s"
        params.append(status)
    if delivery_type:
        query += " AND delivery_type = %s"
        params.append(delivery_type)
        
    query += " ORDER BY timestamp DESC"
    c.execute(query, tuple(params))
    return c.fetchall()

@app.route('/export/csv')
@basic_auth.required
def export_csv():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status')
    delivery_type = request.args.get('delivery_type')
    rows = get_orders_for_export(start_date, end_date, status, delivery_type)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Phone", "Address", "State", "Payment", "Source", "Products", "Total", "Status", "Timestamp", "Notes", "Delivery"])
    for row in rows:
        writer.writerow([row['id'], row['customer_name'], row['phone'], row['address'], row.get('state', ''), row.get('payment_method', 'Prepaid'), row['source'], 
                         row['products'], row['total'], row['status'], row['timestamp'], row['notes'], row.get('delivery_type', 'Standard')])
    
    label = f"{status}_{delivery_type}" if status or delivery_type else "all"
    filename = f"orders_{label}_{start_date or 'all'}_to_{end_date or 'all'}.csv"
    return Response(output.getvalue(), mimetype='text/csv', 
                    headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.route('/export/excel')
@basic_auth.required
def export_excel():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status')
    delivery_type = request.args.get('delivery_type')
    rows = get_orders_for_export(start_date, end_date, status, delivery_type)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Name", "Phone", "Address", "State", "Payment", "Source", "Products", "Total", "Status", "Timestamp", "Notes", "Delivery"])
    for row in rows:
        ws.append([row['id'], row['customer_name'], row['phone'], row['address'], row.get('state', ''), row.get('payment_method', 'Prepaid'), row['source'], 
                   row['products'], row['total'], row['status'], row['timestamp'], row['notes'], row.get('delivery_type', 'Standard')])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    label = f"{status}_{delivery_type}" if status or delivery_type else "all"
    filename = f"orders_{label}_{start_date or 'all'}_to_{end_date or 'all'}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/export/pdf')
@basic_auth.required
def export_pdf():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status')
    delivery_type = request.args.get('delivery_type')
    rows = get_orders_for_export(start_date, end_date, status, delivery_type)

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0, 10, txt="Order Verification Report", ln=True, align='C')
    
    info_txt = f"Date: {start_date or 'All'} to {end_date or 'All'}"
    if status: info_txt += f" | Status: {status}"
    if delivery_type: info_txt += f" | Delivery: {delivery_type}"
    
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 10, txt=info_txt, ln=True, align='C')
    pdf.ln(5)

    # Table Header
    pdf.set_font("Arial", 'B', 10)
    pdf.set_fill_color(240, 240, 240)
    
    # Columns: ID(20), Name(35), Phone(25), State(25), Pay(15), Status(20), Del(20), Total(20), Products(Rest)
    cols = [
        ("ID", 20), ("Name", 35), ("Phone", 25), ("State", 25), ("Pay", 15),
        ("Status", 20), ("Del", 20), ("Total", 20), ("Products", 85)
    ]
    
    for header, width in cols:
        pdf.cell(width, 10, header, border=1, fill=True, align='C')
    pdf.ln()

    # Table Rows
    pdf.set_font("Arial", size=9)
    for row in rows:
        try:
            # Safe Encode
            def clean(text):
                return str(text or "").encode('latin-1', 'replace').decode('latin-1')

            # Truncate logic
            def trunc(text, length=25):
                t = clean(text)
                return t[:length] + "..." if len(t) > length else t

            products_str = clean(row['products'])
            products_str = products_str.replace('[', '').replace(']', '').replace('"', '')

            # Row Height 8
            h = 8
            
            pdf.cell(cols[0][1], h, clean(row['id']), border=1)
            pdf.cell(cols[1][1], h, trunc(row['customer_name'], 18), border=1)
            pdf.cell(cols[2][1], h, clean(row['phone']), border=1)
            pdf.cell(cols[3][1], h, trunc(row.get('state', ''), 12), border=1)
            pdf.cell(cols[4][1], h, clean(row.get('payment_method', 'Prepaid')), border=1)
            pdf.cell(cols[5][1], h, clean(row['status']), border=1)
            pdf.cell(cols[6][1], h, clean(row.get('delivery_type', 'Standard')), border=1)
            pdf.cell(cols[7][1], h, clean(row['total']), border=1)
            pdf.cell(cols[8][1], h, trunc(products_str, 45), border=1)
            
            pdf.ln()
            
        except Exception as e:
            print(f"Error printing PDF row: {e}")
            continue

    output = io.BytesIO()
    output.write(pdf.output(dest='S').encode('latin-1'))
    output.seek(0)
    label = f"{status}_{delivery_type}" if status or delivery_type else "all"
    filename = f"orders_{label}_{start_date or 'all'}_to_{end_date or 'all'}.pdf"
    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)

# --- Webhooks (Public) ---
@app.route('/webhook/shopify', methods=['POST'])
def webhook_shopify():
    order = normalize_shopify_order(request.json)
    if order:
        save_order(order)
    return jsonify({"status": "received"}), 200

@app.route('/webhook/shiprocket', methods=['POST'])
def webhook_shiprocket():
    payload = request.json
    print("=" * 80)
    print("SHIPROCKET WEBHOOK RECEIVED:")
    print(json.dumps(payload, indent=2))
    print("=" * 80)
    
    # Check payment-related fields
    payment_fields = {
        'payment_method': payload.get('payment_method'),
        'cod': payload.get('cod'),
        'is_cod': payload.get('is_cod'),
        'payment_gateway': payload.get('payment_gateway'),
        'payment_mode': payload.get('payment_mode'),
    }
    print("PAYMENT FIELDS FOUND:")
    print(json.dumps(payment_fields, indent=2))
    print("=" * 80)
    
    order = normalize_shiprocket_order(payload)
    if order:
        print(f"NORMALIZED ORDER - Payment Method: {order.get('payment_method')}")
        save_order(order)
    return jsonify({"status": "received"}), 200

# --- Debug (Protected) ---
@app.route('/debug/seed')
@basic_auth.required
def seed_data():
    save_order({
        "id": "#1001", "customer_name": "Amit Sharma", "phone": "+919876543210",
        "email": "amit.sharma@example.com",
        "address": "123, MG Road, Bangalore",
        "state": "Karnataka",
        "payment_method": "Prepaid",
        "rto_risk": "LOW",
        "source": "Shopify", "products": json.dumps(["Blue Shirt - M"]), "total": "1299.00",
        "status": "Pending", "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "notes": "Called once, busy.",
        "delivery_type": "Standard"
    })
    save_order({
        "id": "#5521", "customer_name": "Priya Singh", "phone": "+919988776655",
        "email": "priya.singh@example.com",
        "address": "Green Apts, Mumbai",
        "state": "Maharashtra",
        "payment_method": "COD",
        "rto_risk": "MEDIUM",
        "source": "Shiprocket", "products": json.dumps(["Wireless Earbuds"]), "total": "2499.00",
        "status": "Call Again", "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "notes": "",
        "delivery_type": "Express"
    })
    return redirect(url_for('dashboard'))

# --- Customer Routes ---
@app.route('/customers')
@basic_auth.required
def customers_page():
    """Customer database page"""
    search = request.args.get('search')
    filter_type = request.args.get('filter')
    sort_by = request.args.get('sort', 'last_order_date')
    
    customers = get_all_customers(search=search, filter_type=filter_type, sort_by=sort_by)
    
    # Get customer stats
    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute('''
            SELECT 
                COUNT(*) as total_customers,
                COUNT(*) FILTER (WHERE total_orders > 1) as repeat_customers,
                COALESCE(AVG(total_spent), 0) as avg_lifetime_value
            FROM customers
        ''')
        stats = c.fetchone()
        conn.close()
    except:
        stats = {'total_customers': 0, 'repeat_customers': 0, 'avg_lifetime_value': 0}
    
    return render_template('customers.html', 
                         customers=customers, 
                         stats=stats,
                         search=search,
                         filter_type=filter_type,
                         sort_by=sort_by)

@app.route('/api/customer/<phone>')
@basic_auth.required
def get_customer_api(phone):
    """Get customer profile via API"""
    customer = get_customer_by_phone(phone)
    if customer:
        return jsonify(dict(customer))
    return jsonify({'error': 'Customer not found'}), 404

@app.route('/api/customer/<phone>/orders')
@basic_auth.required
def get_customer_orders(phone):
    """Get all orders for a customer"""
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('SELECT * FROM orders WHERE phone = %s ORDER BY timestamp DESC', (phone,))
    orders = c.fetchall()
    conn.close()
    return jsonify([dict(order) for order in orders])

@app.route('/api/customer/<phone>/notes', methods=['POST'])
@basic_auth.required
def add_customer_note(phone):
    """Add note to customer profile"""
    note = request.json.get('note')
    if not note:
        return jsonify({'error': 'Note required'}), 400
    
    conn = get_db_connection()
    c = conn.cursor()
    
    # Get existing notes
    c.execute('SELECT notes FROM customers WHERE phone = %s', (phone,))
    customer = c.fetchone()
    
    if not customer:
        return jsonify({'error': 'Customer not found'}), 404
    
    # Append new note with timestamp
    existing_notes = customer['notes'] or ''
    timestamp = datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S')
    new_note = f"[{timestamp}] {note}"
    updated_notes = f"{existing_notes}\n{new_note}" if existing_notes else new_note
    
    c.execute('UPDATE customers SET notes = %s WHERE phone = %s', (updated_notes, phone))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(debug=True)
